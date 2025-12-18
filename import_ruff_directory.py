#!/usr/bin/env python3
"""
Script to import contact information from the Ruff Family Directory Excel file
and add/update people in the family tree.
"""

import json
import sys
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import openpyxl


def load_family_data():
    """Load the family tree JSON data."""
    with open('family_tree.json', 'r') as f:
        return json.load(f)


def save_family_data(data):
    """Save the family tree JSON data."""
    with open('family_tree.json', 'w') as f:
        json.dump(data, f, indent=2)
    print(f"\n💾 Saved updated family tree to family_tree.json")


def create_person_id(name: str) -> str:
    """
    Create a person ID from their name.

    Args:
        name: Full name of the person

    Returns:
        ID string (lowercase, underscores instead of spaces)
    """
    # Remove special characters and convert to lowercase
    clean_name = re.sub(r'[^\w\s-]', '', name.lower())
    # Replace spaces and hyphens with underscores
    person_id = re.sub(r'[\s-]+', '_', clean_name)
    return person_id


def parse_name_field(name_field: str) -> Tuple[Optional[str], Optional[str], List[str]]:
    """
    Parse the First Name field which may contain multiple names and relationships.

    Args:
        name_field: The first name field from the directory

    Returns:
        Tuple of (first_name, spouse_name, additional_people)

    Examples:
        "Tom & Annemarie Thompson" -> ("Tom", "Annemarie Thompson", [])
        "Patrick & Jenny Wang" -> ("Patrick", "Jenny Wang", [])
        "Mary Elizabeth" -> ("Mary Elizabeth", None, [])
    """
    # Check for & pattern indicating spouse
    if '&' in name_field:
        parts = name_field.split('&', 1)
        first_name = parts[0].strip()
        spouse_name = parts[1].strip()
        return first_name, spouse_name, []
    else:
        return name_field.strip(), None, []


def parse_phone(phone_str: Optional[str]) -> Optional[str]:
    """
    Parse phone number, handling various formats and extracting primary number.

    Args:
        phone_str: Phone number string (may contain labels like "P" or "M")

    Returns:
        Cleaned phone number or None
    """
    if not phone_str or phone_str == 'N/A':
        return None

    # Remove labels like "302 995-1365 P" -> "302 995-1365"
    phone_str = re.sub(r'\s+[A-Z]$', '', phone_str.strip())

    # If multiple numbers separated by semicolon, take the first
    if ';' in phone_str:
        phone_str = phone_str.split(';')[0].strip()

    return phone_str if phone_str else None


def parse_children(children_str: Optional[str]) -> List[str]:
    """
    Parse the children field which may contain multiple names.

    Args:
        children_str: String containing children names separated by commas

    Returns:
        List of child names
    """
    if not children_str or children_str == 'N/A':
        return []

    # Split by comma and clean up
    children = [child.strip() for child in children_str.split(',')]
    return [c for c in children if c and c != 'N/A']


def parse_address(address: Optional[str], address2: Optional[str],
                  city: Optional[str], state: Optional[str],
                  zip_code: Optional[str]) -> Dict[str, str]:
    """
    Parse address fields into a structured format.

    Returns:
        Dictionary with address components
    """
    result = {}

    if address and address != 'N/A':
        result['address'] = address
    if address2 and address2 != 'N/A':
        result['address2'] = address2
    if city and city != 'N/A':
        result['city'] = city
        result['home_city'] = city  # Also use for home_city field
    if state and state != 'N/A':
        result['state'] = state
        result['home_state'] = state  # Also use for home_state field
    if zip_code and zip_code != 'N/A':
        result['zip'] = str(zip_code)

    return result


def import_directory_row(last_name: str, first_name: str, phone: str,
                        cell: str, email: str, children: str,
                        address: str, address2: str, city: str,
                        state: str, zip_code: str,
                        people: Dict) -> List[Dict]:
    """
    Import a single row from the directory and create person entries.

    Returns:
        List of person dictionaries created from this row
    """
    new_people = []

    # Parse the name field to identify individuals
    first_part, spouse_part, additional = parse_name_field(first_name)

    # Create primary person
    if first_part:
        full_name = f"{first_part} {last_name}".strip()
        person_id = create_person_id(full_name)

        # Check if person already exists
        existing_person = None
        for pid, pdata in people.items():
            if pdata.get('name', '').lower() == full_name.lower():
                existing_person = pid
                break

        person = {
            'id': existing_person or person_id,
            'name': full_name,
            'dob': 'Unknown',
            'dod': 'alive'
        }

        # Add contact information
        phone_num = parse_phone(phone) or parse_phone(cell)
        if phone_num:
            person['phone'] = phone_num

        if email and email != 'N/A':
            person['email'] = email

        # Add address information
        addr_data = parse_address(address, address2, city, state, zip_code)
        person.update(addr_data)

        # Parse children
        child_names = parse_children(children)
        if child_names:
            person['notes'] = f"Children: {', '.join(child_names)}"

        # Add spouse if present
        if spouse_part:
            spouse_name = f"{spouse_part.split()[0] if ' ' in spouse_part else spouse_part} {last_name if ' ' not in spouse_part else spouse_part.split()[-1]}"
            spouse_id = create_person_id(spouse_name)
            person['spouseId'] = spouse_id

            # Create spouse entry
            spouse_person = {
                'id': spouse_id,
                'name': spouse_name,
                'dob': 'Unknown',
                'dod': 'alive',
                'spouseId': existing_person or person_id
            }

            # Check if spouse maiden name is different
            if ' ' in spouse_part and spouse_part.split()[-1] != last_name:
                spouse_person['maidenName'] = spouse_part.split()[-1]

            new_people.append(spouse_person)

        new_people.append(person)

    return new_people


def read_excel_directory(file_path: str) -> List[Dict]:
    """
    Read the Excel directory file and extract contact information.

    Args:
        file_path: Path to the Excel file

    Returns:
        List of person dictionaries
    """
    print(f"📖 Reading Excel file: {file_path}")

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Get header row (assuming first row)
    headers = [cell.value for cell in sheet[1]]
    print(f"📋 Found columns: {headers}")

    # Find column indices
    col_map = {}
    expected_cols = ['Last Name', 'First Name', 'Phone', 'Cell', 'Email',
                     'Children', 'Address', 'Address2', 'City', 'State', 'Zip']

    for i, header in enumerate(headers):
        if header in expected_cols:
            col_map[header] = i

    print(f"✅ Mapped columns: {list(col_map.keys())}")

    # Read data rows
    all_people = []
    row_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not row or not any(row):
            continue

        # Extract values
        last_name = row[col_map.get('Last Name', 0)] or ''
        first_name = row[col_map.get('First Name', 1)] or ''
        phone = row[col_map.get('Phone', 2)] or ''
        cell = row[col_map.get('Cell', 3)] or ''
        email = row[col_map.get('Email', 4)] or ''
        children = row[col_map.get('Children', 5)] or ''
        address = row[col_map.get('Address', 6)] or ''
        address2 = row[col_map.get('Address2', 7)] or ''
        city = row[col_map.get('City', 8)] or ''
        state = row[col_map.get('State', 9)] or ''
        zip_code = row[col_map.get('Zip', 10)] or ''

        # Skip if no name
        if not last_name or not first_name:
            continue

        row_count += 1
        print(f"  Processing row {row_count}: {first_name} {last_name}")

    return all_people


def import_directory(excel_file: str, dry_run: bool = False):
    """
    Import contacts from Excel directory and add to family tree.

    Args:
        excel_file: Path to the Excel file
        dry_run: If True, show what would be added without saving
    """
    print("\n" + "="*80)
    print("📇 RUFF FAMILY DIRECTORY IMPORT")
    print("="*80 + "\n")

    # Load existing family tree
    family_data = load_family_data()
    people = family_data['family']['people']

    print(f"📊 Current family tree has {len(people)} people")

    # Read Excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Get headers
    headers = [cell.value for cell in sheet[1]]

    # Process each row
    new_people = []
    updated_people = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row or not any(row):
            continue

        # Extract values by column name
        row_dict = dict(zip(headers, row))

        last_name = row_dict.get('Last Name', '') or ''
        first_name = row_dict.get('First Name', '') or ''

        if not last_name or not first_name:
            continue

        # Parse and create person entries
        persons = import_directory_row(
            last_name=last_name,
            first_name=first_name,
            phone=row_dict.get('Phone', ''),
            cell=row_dict.get('Cell', ''),
            email=row_dict.get('Email', ''),
            children=row_dict.get('Children', ''),
            address=row_dict.get('Address', ''),
            address2=row_dict.get('Address2', ''),
            city=row_dict.get('City', ''),
            state=row_dict.get('State', ''),
            zip_code=str(row_dict.get('Zip', '') or ''),
            people=people
        )

        # Add or update people
        for person in persons:
            person_id = person['id']

            if person_id in people:
                # Update existing person
                existing = people[person_id]

                # Merge data (prefer new data if present)
                for key, value in person.items():
                    if key != 'id' and value and value not in ['N/A', 'Unknown']:
                        if key not in existing or existing[key] in ['N/A', 'Unknown', '']:
                            existing[key] = value

                updated_people.append(person['name'])
                print(f"  ✏️  Updated: {person['name']}")
            else:
                # Add new person
                people[person_id] = person
                new_people.append(person['name'])
                print(f"  ✨ Added: {person['name']}")

    # Summary
    print("\n" + "="*80)
    print("📊 IMPORT SUMMARY")
    print("="*80)
    print(f"✨ New people added: {len(new_people)}")
    print(f"✏️  People updated: {len(updated_people)}")
    print(f"📊 Total people in tree: {len(people)}")

    if dry_run:
        print("\n🔍 DRY RUN - No changes saved")
        print("\nRun without --dry-run to save changes")
    else:
        # Save updated family tree
        family_data['family']['people'] = people
        save_family_data(family_data)
        print("\n✅ Family tree updated successfully!")


def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description='Import Ruff Family Directory from Excel file'
    )
    parser.add_argument(
        'excel_file',
        nargs='?',
        default='ruffdirectory/RuffDirectory.xlsx',
        help='Path to the Excel directory file (default: ruffdirectory/RuffDirectory.xlsx)'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Show what would be imported without making changes'
    )

    args = parser.parse_args()

    # Check if file exists
    if not Path(args.excel_file).exists():
        print(f"❌ Error: File not found: {args.excel_file}")
        print("\nPlease specify the path to the Excel file:")
        print(f"  python {sys.argv[0]} path/to/directory.xlsx")
        sys.exit(1)

    # Import directory
    import_directory(args.excel_file, dry_run=args.dry_run)


if __name__ == '__main__':
    main()
